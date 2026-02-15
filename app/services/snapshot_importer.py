"""
Snapshot importer service for sqlite backend.

This module validates and imports Excel snapshot files into the
database using a raw sqlite3 connection. It enforces naming
conventions, inserts or updates champions and strategy categories,
populates projects and events tables, and returns a report object.
"""

import re
from datetime import datetime
from typing import List

from fastapi import UploadFile
from openpyxl import load_workbook

from ..db import get_connection
from ..schemas import SnapshotReport


FILENAME_PATTERN = re.compile(r"^(\d{4})-(\d{2})-(\d{2})\.xlsx$")
MONTH_SHEET_PATTERN = re.compile(r"^\d{4}-\d{2}$")


def import_snapshot(file: UploadFile) -> SnapshotReport:
    """Validate and import an Excel snapshot into the SQLite database."""
    warnings: List[str] = []
    errors: List[str] = []

    filename = file.filename or ""
    match = FILENAME_PATTERN.match(filename)
    if not match:
        return SnapshotReport(
            success=False,
            message="Invalid filename format. Must be YYYY-MM-DD.xlsx",
            processed_projects=0,
            processed_events=0,
            warnings=[],
            errors=["Filename does not match pattern"]
        )
    year, month, day = match.groups()
    snapshot_date = f"{year}-{month}-{day}"

    # Load workbook
    try:
        # openpyxl can load from a file-like object directly
        wb = load_workbook(file.file, data_only=True)
    except Exception as e:
        return SnapshotReport(
            success=False,
            message=f"Failed to load workbook: {e}",
            processed_projects=0,
            processed_events=0,
            warnings=[],
            errors=["Workbook load error"]
        )

    if "AX_Master" not in wb.sheetnames:
        return SnapshotReport(
            success=False,
            message="AX_Master sheet is missing.",
            processed_projects=0,
            processed_events=0,
            warnings=[],
            errors=["AX_Master missing"]
        )

    master_ws = wb["AX_Master"]
    master_headers = [cell.value for cell in next(master_ws.iter_rows(min_row=1, max_row=1))]
    expected_master_cols = {
        "과제ID": "project_id",
        "과제명": "project_name",
        "Champion": "champion",
        "전략분류": "strategy",
        "수행 부서": "org_unit",
        "심의상태": "status",
        "제안월": "proposed_month",
        "승인월": "approved_month",
    }
    header_index = {}
    for idx, header in enumerate(master_headers):
        if header in expected_master_cols:
            header_index[expected_master_cols[header]] = idx
    missing = [k for k, v in expected_master_cols.items() if v not in header_index]
    if missing:
        return SnapshotReport(
            success=False,
            message=f"Missing required columns in AX_Master: {', '.join(missing)}",
            processed_projects=0,
            processed_events=0,
            warnings=[],
            errors=["Missing columns"]
        )

    processed_projects = 0
    processed_events = 0

    conn = get_connection()
    c = conn.cursor()
    try:
        # Check duplicate snapshot date
        c.execute("SELECT snapshot_id FROM snapshots WHERE snapshot_date = ?", (snapshot_date,))
        if c.fetchone():
            return SnapshotReport(
                success=False,
                message="Snapshot for this date already exists.",
                processed_projects=0,
                processed_events=0,
                warnings=[],
                errors=["Duplicate snapshot date"]
            )
        # Insert snapshot
        uploaded_at = datetime.utcnow().isoformat()
        c.execute(
            "INSERT INTO snapshots (snapshot_date, uploaded_at, source_filename) VALUES (?,?,?)",
            (snapshot_date, uploaded_at, filename)
        )
        snapshot_id = c.lastrowid
        # Function to get or create champion
        def get_champion_id(name: str):
            if not name:
                return None
            name = name.strip()
            if not name:
                return None
            c.execute("SELECT champion_id FROM champions WHERE name = ?", (name,))
            row = c.fetchone()
            if row:
                return row[0]
            c.execute("INSERT INTO champions (name) VALUES (?)", (name,))
            return c.lastrowid
        # Function to get or create strategy
        def get_strategy_id(name: str):
            if not name:
                return None
            name = name.strip()
            if not name:
                return None
            c.execute("SELECT strategy_id FROM strategy_categories WHERE name = ?", (name,))
            row = c.fetchone()
            if row:
                return row[0]
            c.execute("INSERT INTO strategy_categories (name) VALUES (?)", (name,))
            return c.lastrowid
        # Process AX_Master rows
        for row in master_ws.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue
            project_id = row[header_index["project_id"]].value
            if not project_id:
                warnings.append("Blank project_id in AX_Master; row skipped")
                continue
            project_name = row[header_index["project_name"]].value or ""
            champion_name = row[header_index["champion"]].value
            strategy_name = row[header_index["strategy"]].value
            org_unit = row[header_index["org_unit"]].value
            status = row[header_index["status"]].value or "제안"
            proposed_month = row[header_index["proposed_month"]].value
            approved_month = row[header_index["approved_month"]].value
            champion_id = get_champion_id(str(champion_name)) if champion_name else None
            strategy_id = get_strategy_id(str(strategy_name)) if strategy_name else None
            c.execute(
                """
                INSERT INTO projects (
                    snapshot_id, project_id, project_name, champion_id, strategy_id,
                    org_unit, current_status, proposed_month, approved_month
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    str(project_id),
                    str(project_name),
                    champion_id,
                    strategy_id,
                    str(org_unit) if org_unit else None,
                    str(status),
                    str(proposed_month) if proposed_month else None,
                    str(approved_month) if approved_month else None,
                )
            )
            processed_projects += 1
        # Process monthly sheets
        for sheet_name in wb.sheetnames:
            if sheet_name == "AX_Master":
                continue
            if not MONTH_SHEET_PATTERN.match(sheet_name):
                warnings.append(f"Sheet {sheet_name} ignored due to invalid name")
                continue
            month_key = sheet_name
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            expected_event_cols = {
                "과제ID": "project_id",
                "Champion": "champion",
                "신규제안여부": "is_new_proposal",
                "승인여부": "is_approved",
                "비고": "note",
            }
            event_index = {}
            for idx, h in enumerate(headers):
                if h in expected_event_cols:
                    event_index[expected_event_cols[h]] = idx
            missing_evt_cols = [k for k in expected_event_cols.values() if k not in event_index]
            if missing_evt_cols:
                conn.rollback()
                return SnapshotReport(
                    success=False,
                    message=f"Missing required columns in sheet {sheet_name}:",
                    processed_projects=processed_projects,
                    processed_events=processed_events,
                    warnings=warnings,
                    errors=["Missing event columns"]
                )
            # Parse events rows
            for row in ws.iter_rows(min_row=2):
                if all(cell.value is None for cell in row):
                    continue
                p_id = row[event_index["project_id"]].value
                if not p_id:
                    warnings.append(f"Blank project_id in {sheet_name}; row skipped")
                    continue
                # Check project exists
                c.execute("SELECT 1 FROM projects WHERE snapshot_id = ? AND project_id = ?", (snapshot_id, str(p_id)))
                if not c.fetchone():
                    conn.rollback()
                    return SnapshotReport(
                        success=False,
                        message=f"Project ID {p_id} in sheet {sheet_name} not found in AX_Master.",
                        processed_projects=processed_projects,
                        processed_events=processed_events,
                        warnings=warnings,
                        errors=[f"Unknown project_id {p_id}"]
                    )
                champ_name = row[event_index["champion"]].value
                champ_id = None
                if champ_name and str(champ_name).strip():
                    champ_id = get_champion_id(str(champ_name))
                else:
                    # Use project's champion
                    c.execute("SELECT champion_id FROM projects WHERE snapshot_id = ? AND project_id = ?", (snapshot_id, str(p_id)))
                    proj_row = c.fetchone()
                    champ_id = proj_row[0] if proj_row else None
                val_new = row[event_index["is_new_proposal"]].value
                is_new = 0
                if val_new is not None and str(val_new).strip() != "0":
                    is_new = 1
                val_app = row[event_index["is_approved"]].value
                is_approved = 0
                if val_app is not None and str(val_app).strip() != "0":
                    is_approved = 1
                note = row[event_index["note"]].value
                c.execute(
                    """
                    INSERT INTO project_monthly_events (
                        snapshot_id, month_key, project_id, champion_id, is_new_proposal, is_approved, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        snapshot_id,
                        month_key,
                        str(p_id),
                        champ_id,
                        is_new,
                        is_approved,
                        str(note) if note else None,
                    )
                )
                processed_events += 1
        conn.commit()
    except Exception as e:
        conn.rollback()
        return SnapshotReport(
            success=False,
            message=f"Error during import: {e}",
            processed_projects=processed_projects,
            processed_events=processed_events,
            warnings=warnings,
            errors=["Unhandled exception during import"]
        )
    finally:
        conn.close()
    return SnapshotReport(
        success=True,
        message="Snapshot imported successfully.",
        processed_projects=processed_projects,
        processed_events=processed_events,
        warnings=warnings,
        errors=errors,
    )